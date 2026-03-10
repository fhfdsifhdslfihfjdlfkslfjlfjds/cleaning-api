from fastapi import FastAPI, UploadFile, File, Form, HTTPException
from fastapi.responses import FileResponse
from typing import Optional, Any
from tempfile import NamedTemporaryFile
from datetime import datetime, timedelta, date
from openpyxl.styles import PatternFill
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import tempfile
import os
import re
import openpyxl
import zipfile
import shutil
from pathlib import Path


app = FastAPI(title="Cleaning Schedule API", version="0.1.0")

DEFAULT_ORDER_SHEET_NAME = "担当者割当順番シート"
DEFAULT_STATE_SHEET_NAME = "状態管理シート"
VERSION_SHEET_NAME = "版管理"
MASTER_ID = "cleaning_master"
YELLOW_RGB_SET = {"FFFFFF00", "FFFF00", "00FFFF00"}
YELLOW_FILL = PatternFill(fill_type="solid", fgColor="FFFF00")


# =========================
# 共通ヘルパー
# =========================
def _cell_str(v: Any) -> str:
    if v is None:
        return ""
    return str(v).strip()

def normalize_person_name(name: str) -> str:
    """
    シフト表の氏名から照合用の苗字を取り出す
    例:
    - '服部 翔　(7時)' -> '服部'
    - '田中 達也 (8時)' -> '田中'
    - '栄 真吾' -> '栄'
    """
    s = _cell_str(name)

    # 全角スペースを半角スペースに寄せる
    s = s.replace("　", " ")

    # カッコ以降を削除
    s = re.sub(r"\(.*?\)", "", s).strip()

    # 空白区切りの先頭だけ返す
    parts = [p for p in s.split(" ") if p]
    if parts:
        return parts[0]

    return s


def parse_pipe_list(value: Any) -> list[str]:
    s = _cell_str(value)
    if not s:
        return []
    return [x.strip() for x in s.split("|") if x.strip()]


def join_pipe_list(items: list[str]) -> str:
    return "|".join([x for x in items if _cell_str(x)])


def unique_keep_order(items: list[str]) -> list[str]:
    seen = set()
    result = []
    for x in items:
        if x not in seen:
            seen.add(x)
            result.append(x)
    return result


def normalize_place_name(name: str) -> str:
    s = _cell_str(name)
    if "事務所前" in s:
        return "事務所前"
    if "男子" in s:
        return "男子トイレ"
    if "女子" in s:
        return "女子トイレ"
    return s


def is_yellow_fill(cell) -> bool:
    fill = cell.fill
    if not fill or not fill.fill_type:
        return False

    rgb = None
    if fill.fgColor and fill.fgColor.rgb:
        rgb = fill.fgColor.rgb
    elif fill.start_color and fill.start_color.rgb:
        rgb = fill.start_color.rgb

    if not rgb:
        return False

    return rgb.upper() in YELLOW_RGB_SET


def clear_yellow_fill(cell) -> None:
    cell.fill = PatternFill(fill_type=None)


# =========================
# シフト表解析
# =========================
def get_start_date_from_sheet_name(sheet_name: str) -> date:
    """
    例: 20260316-20260415 -> 2026-03-16
    """
    m = re.match(r"(\d{8})-\d{8}", sheet_name)
    if not m:
        raise ValueError(f"シート名 '{sheet_name}' から開始日を取得できません")
    return datetime.strptime(m.group(1), "%Y%m%d").date()


def find_target_header_row(ws, scan_rows: int = 120, scan_cols: int = 80):
    """
    下側の日付行を見つける
    条件:
    - 1〜31 の数字が横に7個以上並ぶ
    - 次の行に曜日が複数ある
    - さらにその下にA列の名前がある
    条件を満たす候補のうち、一番下を採用
    """
    weekday_set = {"月", "火", "水", "木", "金", "土", "日"}
    candidates = []

    for r in range(1, min(scan_rows, ws.max_row) + 1):
        day_cols = []
        for c in range(1, min(scan_cols, ws.max_column) + 1):
            s = _cell_str(ws.cell(r, c).value)
            if s.isdigit():
                d = int(s)
                if 1 <= d <= 31:
                    day_cols.append((c, d))

        if len(day_cols) < 7:
            continue

        weekday_count = 0
        for c, _ in day_cols:
            s = _cell_str(ws.cell(r + 1, c).value)
            if s in weekday_set:
                weekday_count += 1

        if weekday_count < 5:
            continue

        name_found = False
        for rr in range(r + 2, min(r + 12, ws.max_row) + 1):
            name = _cell_str(ws.cell(rr, 1).value)
            if name:
                name_found = True
                break

        if not name_found:
            continue

        candidates.append((r, day_cols))

    if not candidates:
        raise ValueError("対象の日付ヘッダー行が見つかりませんでした")

    header_row, day_cols = candidates[-1]
    day_cols.sort(key=lambda x: x[0])
    return header_row, day_cols


def build_date_info_from_start_date(start_date: date, day_cols, weekday_row_values=None):
    """
    左から順に start_date + idx 日として実日付を付与
    """
    date_info = []
    for idx, (col_idx, day_num) in enumerate(day_cols):
        actual_date = start_date + timedelta(days=idx)
        weekday_char = ""
        if weekday_row_values and col_idx in weekday_row_values:
            weekday_char = weekday_row_values[col_idx]

        date_info.append({
            "col": col_idx,
            "day": day_num,
            "date": actual_date.isoformat(),
            "weekday": weekday_char,
        })
    return date_info


def parse_shift_table_for_cleaning(ws, sheet_name: str):
    """
    シフト表専用ロジック
    - 下側の日付行を使う
    - A列を名前列とする
    - 曜日行の次以降でA列に名前が出る最初の行から読む
    - 空白行は1回だけ許容、2回連続で終了
    - (定時) 行や、日付列が全部空の行は除外
    """
    header_row, day_cols = find_target_header_row(ws)

    weekday_row = header_row + 1
    weekday_row_values = {}
    for col_idx, _ in day_cols:
        weekday_row_values[col_idx] = _cell_str(ws.cell(weekday_row, col_idx).value)

    start_date = get_start_date_from_sheet_name(sheet_name)
    date_info = build_date_info_from_start_date(start_date, day_cols, weekday_row_values)

    data_start_row = None
    for r in range(weekday_row + 1, min(weekday_row + 15, ws.max_row) + 1):
        name = _cell_str(ws.cell(r, 1).value)
        if name:
            data_start_row = r
            break

    if data_start_row is None:
        raise ValueError("メンバー行が見つかりませんでした")

    people = []
    blank_name_streak = 0

    for r in range(data_start_row, ws.max_row + 1):
        name = _cell_str(ws.cell(r, 1).value)

        if not name:
            blank_name_streak += 1
            if blank_name_streak >= 2:
                break
            continue

        blank_name_streak = 0

        if name in {"氏名", "名前", "合計", "小計"}:
            continue
        if "(定時)" in name:
            continue

        cells = {}
        non_empty_count = 0

        for d in date_info:
            col_idx = d["col"]
            v = _cell_str(ws.cell(r, col_idx).value)
            cells[d["date"]] = v
            if v != "":
                non_empty_count += 1

        # 日付列が全部空ならメンバー行ではない
        if non_empty_count == 0:
            continue

        people.append({
            "row": r,
            "name": name,  # 元の表示名
            "normalized_name": normalize_person_name(name),  # 照合用の苗字
            "cells": cells,
        })

    return {
        "header_row": header_row,
        "weekday_row": weekday_row,
        "data_start_row": data_start_row,
        "date_info": date_info,
        "people": people,
    }


# =========================
# 順番表シート解析
# =========================
def parse_order_sheet(ws):
    """
    1行目: 場所名
    2行目以降: 順番
    黄色セル: 初回開始位置
    """
    places = {}

    for col in range(1, ws.max_column + 1):
        raw_place_name = _cell_str(ws.cell(1, col).value)
        if not raw_place_name:
            continue

        place_name = normalize_place_name(raw_place_name)

        members = []
        start_from = None
        blank_streak = 0

        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row, col)
            name = _cell_str(cell.value)

            if not name:
                blank_streak += 1
                if blank_streak >= 2:
                    break
                continue

            blank_streak = 0
            normalized = normalize_person_name(name)
            members.append(normalized)

            if is_yellow_fill(cell):
                start_from = normalized

        if members:
            if start_from is None:
                start_from = members[0]

            places[place_name] = {
                "members": members,
                "start_from": start_from,
                "col": col,
            }

    return places


# =========================
# 状態管理シート
# =========================
def ensure_state_sheet(wb, sheet_name: str, place_names):
    if sheet_name not in wb.sheetnames:
        ws = wb.create_sheet(sheet_name)
        ws["A1"] = "場所"
        ws["B1"] = "次回先頭"
        ws["C1"] = "保留リスト"
        ws["D1"] = "前回担当リスト"
        ws["E1"] = "前回担当日"
        ws["F1"] = "スキップリスト"
        ws["G1"] = "更新日時"
        ws["H1"] = "メモ"

        row = 2
        for place in place_names:
            ws.cell(row=row, column=1, value=place)
            row += 1
    else:
        ws = wb[sheet_name]

    existing_places = set()
    for row in range(2, ws.max_row + 1):
        p = _cell_str(ws.cell(row, 1).value)
        if p:
            existing_places.add(p)

    next_row = ws.max_row + 1
    for place in place_names:
        if place not in existing_places:
            ws.cell(row=next_row, column=1, value=place)
            next_row += 1

    return ws


def read_state_sheet(ws, order_data):
    """
    空欄なら順番表シートの黄色セルを初期開始位置として使う
    """
    state = {}

    for row in range(2, ws.max_row + 1):
        place = _cell_str(ws.cell(row, 1).value)
        if not place:
            continue

        next_start = _cell_str(ws.cell(row, 2).value)
        carry_list = parse_pipe_list(ws.cell(row, 3).value)
        prev_assigned_list = parse_pipe_list(ws.cell(row, 4).value)
        prev_assigned_date = _cell_str(ws.cell(row, 5).value)
        skip_list = parse_pipe_list(ws.cell(row, 6).value)

        if not next_start and place in order_data:
            next_start = order_data[place]["start_from"]

        state[place] = {
            "row": row,
            "next_start": next_start,
            "carry_list": carry_list,
            "prev_assigned_list": prev_assigned_list,
            "prev_assigned_date": prev_assigned_date,
            "skip_list": skip_list,
        }

    for place, info in order_data.items():
        if place not in state:
            state[place] = {
                "row": None,
                "next_start": info["start_from"],
                "carry_list": [],
                "prev_assigned_list": [],
                "prev_assigned_date": "",
                "skip_list": [],
            }

    return state


def write_state_sheet(ws, state):
    row_map = {}
    for row in range(2, ws.max_row + 1):
        place = _cell_str(ws.cell(row, 1).value)
        if place:
            row_map[place] = row

    next_row = ws.max_row + 1

    for place, info in state.items():
        if place in row_map:
            row = row_map[place]
        else:
            row = next_row
            ws.cell(row=row, column=1, value=place)
            next_row += 1

        ws.cell(row=row, column=2, value=info.get("next_start", ""))
        ws.cell(row=row, column=3, value=join_pipe_list(info.get("carry_list", [])))
        ws.cell(row=row, column=4, value=join_pipe_list(info.get("prev_assigned_list", [])))
        ws.cell(row=row, column=5, value=info.get("prev_assigned_date", ""))
        ws.cell(row=row, column=6, value=join_pipe_list(info.get("skip_list", [])))
        ws.cell(row=row, column=7, value=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))


# =========================
# 割当ロジック
# =========================
def rotate_list_from_name(members, start_from):
    if not members:
        return []
    if not start_from or start_from not in members:
        return members[:]
    idx = members.index(start_from)
    return members[idx:] + members[:idx]


def is_working_on_date(people_map, name, date_str):
    """
    空白なら出勤、文字ありなら休み
    """
    person = people_map.get(name)
    if not person:
        return False

    v = person["cells"].get(date_str, "")
    return _cell_str(v) == ""


def assign_cleaning_with_state(parsed_shift, order_data, state_data, clean_days_normalized):
    """
    - 場所ごと独立
    - carry_list を先頭優先
    - skip_list は次回1回だけ飛ばす
    - 同じ日に同じ人が複数場所担当でもOK
    - 同じ周回で一度担当した人は再登場しない
    - ただし、その周回の未担当者が全員休みなら、その日の中で次周回へ入る
    - carry から担当が出ても、carry の後ろに残る人は保持する
    """
    target_weekdays = set(clean_days_normalized)

    people_map = {}
    for p in parsed_shift["people"]:
        key = p.get("normalized_name", p["name"])
        people_map[key] = p

    target_dates = [
        d for d in parsed_shift["date_info"]
        if d["weekday"] in target_weekdays
    ]

    runtime = {}
    for place, order_info in order_data.items():
        st = state_data.get(place, {})
        runtime[place] = {
            "members": order_info["members"][:],
            "current_start": st.get("next_start") or order_info["start_from"],
            "carry_list": st.get("carry_list", [])[:],
            "skip_once": st.get("skip_list", [])[:],
            "assigned_history": [],
            "last_assigned_date": "",
            "used_in_cycle": set(),
        }

    def pick_from_candidates(rt, carry_candidates, regular_candidates, date_str):
        """
        carry → regular の順で探索
        戻り値:
        - assigned
        - assigned_source ("carry" / "regular" / None)
        - carry_after （次回へ残す carry）
        """
        carry_skipped = []

        # 1) carry を探索
        for idx, name in enumerate(carry_candidates):
            if name in rt["skip_once"]:
                rt["skip_once"].remove(name)
                continue

            if is_working_on_date(people_map, name, date_str):
                # carry から担当が出た場合、
                # 後ろに残っている carry は次回へ残す
                remaining_carry = carry_candidates[idx + 1:]
                carry_after = unique_keep_order(carry_skipped + remaining_carry)
                return name, "carry", carry_after
            else:
                carry_skipped.append(name)

        # 2) regular を探索
        regular_skipped = []
        for idx, name in enumerate(regular_candidates):
            if name in rt["skip_once"]:
                rt["skip_once"].remove(name)
                continue

            if is_working_on_date(people_map, name, date_str):
                # regular から担当が出た場合、
                # それまで休みで飛ばした regular は carry に入れる
                carry_after = unique_keep_order(carry_skipped + regular_skipped)
                return name, "regular", carry_after
            else:
                regular_skipped.append(name)

        # 誰も担当できなかった
        carry_after = unique_keep_order(carry_skipped + regular_skipped)
        return None, None, carry_after

    results = []

    for d in target_dates:
        date_str = d["date"]
        weekday = d["weekday"]

        day_result = {
            "date": date_str,
            "weekday": weekday,
            "assignments": {}
        }

        for place, rt in runtime.items():
            for place, rt in runtime.items():
            members = rt["members"]
            rotated = rotate_list_from_name(members, rt["current_start"])

            # DEBUG①
            if place == "事務所前" and date_str == "2026-03-27":
                print("DEBUG 事務所前 2026-03-27")
                print("current_start =", rt["current_start"])
                print("carry_list =", rt["carry_list"])
                print("skip_once =", rt["skip_once"])
                print("used_in_cycle =", rt["used_in_cycle"])
                print("rotated =", rotated)

            # carry と regular を分離
            base_carry = unique_keep_order(rt["carry_list"])
            base_regular = [x for x in rotated if x not in base_carry]

            # -----------------------------
            # 1回目: この周回で未担当の人だけで探索
            # -----------------------------
            first_carry = [x for x in base_carry if x not in rt["used_in_cycle"]]
            first_regular = [x for x in base_regular if x not in rt["used_in_cycle"]]

            # DEBUG②
            if place == "事務所前" and date_str == "2026-03-27":
                print("三宅 working? ", is_working_on_date(people_map, "三宅", date_str))
                print("波多野 working? ", is_working_on_date(people_map, "波多野", date_str))
                print("first_regular =", first_regular)

            assigned, assigned_source, carry_after = pick_from_candidates(
                rt, first_carry, first_regular, date_str
            )

            started_new_cycle = False

            # -----------------------------
            # 2回目: 1回目で決まらなければ、その日の中で次周回へ入る
            # -----------------------------
            if not assigned:
                second_carry = [x for x in base_carry if x in rt["used_in_cycle"]]
                second_regular = [x for x in base_regular if x in rt["used_in_cycle"]]

                assigned2, assigned_source2, carry_after2 = pick_from_candidates(
                    rt, second_carry, second_regular, date_str
                )

                if assigned2:
                    assigned = assigned2
                    assigned_source = assigned_source2
                    carry_after = unique_keep_order(carry_after + carry_after2)
                    started_new_cycle = True

            day_result["assignments"][place] = assigned if assigned else ""

            # carry 更新
            rt["carry_list"] = unique_keep_order(carry_after)

            if assigned:
                rt["assigned_history"].append(assigned)
                rt["last_assigned_date"] = date_str

                # 新しい周回に入ったなら used をリセットして今回担当だけ入れる
                if started_new_cycle:
                    rt["used_in_cycle"] = {assigned}
                else:
                    rt["used_in_cycle"].add(assigned)

                # 通常順から選ばれたときだけ current_start を進める
                # carry から復帰した人では current_start を動かさない
                if assigned_source == "regular":
                    idx = members.index(assigned)
                    next_idx = (idx + 1) % len(members)
                    rt["current_start"] = members[next_idx]

        results.append(day_result)

    new_state = {}
    for place, rt in runtime.items():
        if rt["carry_list"]:
            next_start = rt["carry_list"][0]
        else:
            next_start = rt["current_start"]

        assigned_history = rt.get("assigned_history", [])
        last_person = str(assigned_history[-1]) if assigned_history else ""
        prev_list = [last_person] if last_person else []

        new_state[place] = {
            "next_start": next_start,
            "carry_list": rt["carry_list"],
            "prev_assigned_list": prev_list,
            "prev_assigned_date": rt["last_assigned_date"],
            "skip_list": prev_list
        }

    return results, new_state


# =========================
# 順番表シートの黄色更新
# =========================
def update_order_sheet_yellow(order_ws, order_data, new_state):
    """
    順番表シートの黄色セルを、次回先頭の人に更新
    """
    # 既存の黄色を消す
    for place, info in order_data.items():
        col = info["col"]
        for row in range(2, order_ws.max_row + 1):
            clear_yellow_fill(order_ws.cell(row, col))

    # 次回先頭に黄色を付ける
    for place, info in order_data.items():
        col = info["col"]
        next_start = new_state.get(place, {}).get("next_start", "")
        if not next_start:
            continue

        for row in range(2, order_ws.max_row + 1):
            cell = order_ws.cell(row, col)
            if _cell_str(cell.value) == next_start:
                cell.fill = YELLOW_FILL
                break

def apply_box_border(ws, start_row, end_row, start_col, end_col, thin, medium):
    for r in range(start_row, end_row + 1):
        for c in range(start_col, end_col + 1):
            left_side = medium if c == start_col else thin
            right_side = medium if c == end_col else thin
            top_side = medium if r == start_row else thin
            bottom_side = medium if r == end_row else thin

            ws.cell(r, c).border = Border(
                left=left_side,
                right=right_side,
                top=top_side,
                bottom=bottom_side
            )

# =========================
# 出力シート作成
# =========================
def write_cleaning_sheet(wb, sheet_name, assignments, period_label):
    """
    添付画像風の掃除当番表を新規作成
    """
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]

    ws = wb.create_sheet(sheet_name)

    # -------------------------
    # スタイル
    # -------------------------
    thin = Side(style="thin", color="000000")
    medium = Side(style="medium", color="000000")

    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    title_font = Font(size=22, bold=True)
    block_font = Font(size=20, bold=True)
    header_font = Font(size=12, bold=True)
    normal_font = Font(size=11)
    name_font = Font(size=12, bold=True)
    note_font = Font(size=11)

    yellow_fill = PatternFill(fill_type="solid", fgColor="FFF200")
    blue_fill = PatternFill(fill_type="solid", fgColor="9FD5F3")
    pink_fill = PatternFill(fill_type="solid", fgColor="E8C3E8")

    # -------------------------
    # 列幅
    # -------------------------
    widths = {
        "A": 18,
        "B": 13,
        "C": 13,
        "D": 13,
        "E": 13,
        "F": 13,
        "G": 13,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    # -------------------------
    # 上部ヘッダー（1行）
    # -------------------------
    ws.merge_cells("A1:C1")
    ws["A1"] = "掃除当番表"
    ws["A1"].font = Font(size=24, bold=True)
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws["A1"].border = Border()

    ws["E1"] = "対象期間"
    ws["E1"].font = Font(size=12, bold=True)
    ws["E1"].alignment = center
    ws["E1"].border = Border(left=medium, right=thin, top=medium, bottom=medium)

    ws.merge_cells("F1:G1")
    ws["F1"] = period_label
    ws["F1"].font = Font(size=12, bold=True)
    ws["F1"].alignment = center
    ws["F1"].border = Border(left=thin, right=medium, top=medium, bottom=medium)

    # 2行目は空けるだけ
    for cell_ref in ["A2", "B2", "C2", "D2", "E2", "F2", "G2"]:
        ws[cell_ref].value = ""
        ws[cell_ref].border = Border()
        ws[cell_ref].fill = PatternFill(fill_type=None)
    
    ws["E1"].border = Border(left=medium, right=thin, top=medium, bottom=medium)
    ws["F1"].border = Border(left=thin, right=medium, top=medium, bottom=medium)
    ws["G1"].border = Border(left=thin, right=medium, top=medium, bottom=medium)

    # 場所名
    ws.merge_cells("B4:C5")
    ws["B4"] = "1階5℃庫前～\n事務所前"
    ws["B4"].font = header_font
    ws["B4"].alignment = center
    ws["B4"].fill = yellow_fill
    ws["B4"].border = Border(left=medium, right=medium, top=medium, bottom=medium)

    ws.merge_cells("D4:E5")
    ws["D4"] = "男子トイレ"
    ws["D4"].font = block_font
    ws["D4"].alignment = center
    ws["D4"].fill = blue_fill
    ws["D4"].border = Border(left=medium, right=medium, top=medium, bottom=medium)

    ws.merge_cells("F4:G5")
    ws["F4"] = "女子トイレ"
    ws["F4"].font = block_font
    ws["F4"].alignment = center
    ws["F4"].fill = pink_fill
    ws["F4"].border = Border(left=medium, right=medium, top=medium, bottom=medium)

    apply_box_border(ws, 4, 5, 2, 3, thin, medium)  # B4:C5
    apply_box_border(ws, 4, 5, 4, 5, thin, medium)  # D4:E5
    apply_box_border(ws, 4, 5, 6, 7, thin, medium)  # F4:G5

    # 見出し
    ws["A6"] = "日付"
    ws["B6"] = "担当者"
    ws["C6"] = "チェック"
    ws["D6"] = "担当者"
    ws["E6"] = "チェック"
    ws["F6"] = "担当者"
    ws["G6"] = "チェック"

    for cell_ref in ["A6", "B6", "C6", "D6", "E6", "F6", "G6"]:
        ws[cell_ref].font = header_font
        ws[cell_ref].alignment = center
        ws[cell_ref].border = Border(left=medium, right=medium, top=medium, bottom=medium)
    
    for cell_ref in ["F4", "G4", "F5", "G5"]:
        ws[cell_ref].border = Border(left=medium, right=medium, top=medium, bottom=medium)

    # 行高さ
    ws.row_dimensions[1].height = 34
    ws.row_dimensions[2].height = 12
    ws.row_dimensions[4].height = 28
    ws.row_dimensions[5].height = 32
    ws.row_dimensions[6].height = 24

    # -------------------------
    # データ部
    # -------------------------
    start_row = 7

    m = re.match(r"(\d{4})/(\d{1,2})/(\d{1,2})〜(\d{4})/(\d{1,2})/(\d{1,2})", period_label)
    if not m:
        raise ValueError(f"period_label の形式が想定外です: {period_label}")

    start_date = date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
    end_date = date(int(m.group(4)), int(m.group(5)), int(m.group(6)))

    assignment_map = {a["date"]: a["assignments"] for a in assignments}
    weekday_map = {
        0: "月", 1: "火", 2: "水", 3: "木", 4: "金", 5: "土", 6: "日"
    }

    current = start_date
    row = start_row
    while current <= end_date:
        weekday_char = weekday_map[current.weekday()]
        date_label = f"{current.month}/{current.day}({weekday_char})"

        ws.cell(row=row, column=1, value=date_label)
        ws.cell(row=row, column=1).alignment = center
        ws.cell(row=row, column=1).font = normal_font

        assignments_for_day = assignment_map.get(current.isoformat(), {})

        office_name = assignments_for_day.get("事務所前", "")
        men_name = assignments_for_day.get("男子トイレ", "")
        women_name = assignments_for_day.get("女子トイレ", "")

        ws.cell(row=row, column=2, value=office_name)
        ws.cell(row=row, column=4, value=men_name)
        ws.cell(row=row, column=6, value=women_name)

        for col in [2, 4, 6]:
            ws.cell(row=row, column=col).alignment = center
            ws.cell(row=row, column=col).font = name_font

        for col in [3, 5, 7]:
            ws.cell(row=row, column=col, value="")
            ws.cell(row=row, column=col).alignment = center

        # 罫線
        for col in range(1, 8):
            cell = ws.cell(row=row, column=col)

            left_side = medium if col in [1, 2, 4, 6] else thin
            right_side = medium if col in [1, 3, 5, 7] else thin

            top_side = thin
            bottom_side = thin

            if current == end_date:
                bottom_side = medium

            cell.border = Border(
                left=left_side,
                right=right_side,
                top=top_side,
                bottom=bottom_side
            )

        ws.row_dimensions[row].height = 24
        row += 1
        current += timedelta(days=1)

    last_row = row - 1

    # -------------------------
    # 下部メモ
    # -------------------------
    # 注意書きは横に広く使って見切れ防止
    ws.merge_cells("A39:E39")
    ws.merge_cells("A40:E40")

    ws["A39"] = "※清掃終了後、チェック欄に〇を記入"
    ws["A40"] = "※確認者は抜けが無いかチェックし、確認者欄に氏名を記入"

    ws["A39"].font = note_font
    ws["A40"].font = note_font

    ws["A39"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws["A40"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # 行高さを十分に確保
    ws.row_dimensions[39].height = 24
    ws.row_dimensions[40].height = 42

    # 確認者欄
    ws["F40"] = "確認者"
    ws["F40"].font = header_font
    ws["F40"].alignment = center
    ws["F40"].border = Border(bottom=medium)

    ws["G40"] = ""
    ws["G40"].border = Border(bottom=medium)

    # -------------------------
    # 印刷設定
    # -------------------------
    ws.print_area = f"A1:G40"
    ws.print_title_rows = "1:6"

    # A4 1枚に収める
    ws.page_setup.paperSize = 9   # A4
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    # 余白を少し詰める
    ws.page_margins.left = 0.2
    ws.page_margins.right = 0.2
    ws.page_margins.top = 0.3
    ws.page_margins.bottom = 0.3
    ws.page_margins.header = 0.1
    ws.page_margins.footer = 0.1

    # 印刷時に中央寄せ
    ws.print_options.horizontalCentered = True

    return ws

def ensure_version_sheet(wb):
    if VERSION_SHEET_NAME not in wb.sheetnames:
        ws = wb.create_sheet(VERSION_SHEET_NAME)
        ws["A1"] = "master_id"
        ws["B1"] = MASTER_ID
        ws["A2"] = "version"
        ws["B2"] = ""
        ws["A3"] = "previous_version"
        ws["B3"] = ""
        ws["A4"] = "generated_at"
        ws["B4"] = ""
        ws["A5"] = "status"
        ws["B5"] = "current"
    else:
        ws = wb[VERSION_SHEET_NAME]
    return ws


def validate_or_initialize_version_sheet(wb):
    """
    版管理シートがなければ初回作成。
    あれば status/version を検証。
    """
    if VERSION_SHEET_NAME not in wb.sheetnames:
        ws = ensure_version_sheet(wb)
        initial_version = datetime.now().strftime("%Y%m%d_%H%M%S")
        ws["B2"] = initial_version
        ws["B4"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ws["B5"] = "current"
        return {
            "master_id": MASTER_ID,
            "version": initial_version,
            "previous_version": "",
            "generated_at": ws["B4"].value,
            "status": "current",
            "initialized": True,
        }

    ws = wb[VERSION_SHEET_NAME]

    master_id = _cell_str(ws["B1"].value)
    version = _cell_str(ws["B2"].value)
    previous_version = _cell_str(ws["B3"].value)
    generated_at = _cell_str(ws["B4"].value)
    status = _cell_str(ws["B5"].value)

    if master_id != MASTER_ID:
        raise HTTPException(status_code=400, detail="マスタファイルの master_id が不正です")

    if not version:
        raise HTTPException(status_code=400, detail="マスタファイルの version が未設定です")

    if status != "current":
        raise HTTPException(status_code=400, detail="このマスタファイルは最新版ではありません")

    return {
        "master_id": master_id,
        "version": version,
        "previous_version": previous_version,
        "generated_at": generated_at,
        "status": status,
        "initialized": False,
    }


def update_version_info(wb):
    ws = ensure_version_sheet(wb)

    old_version = _cell_str(ws["B2"].value)
    new_version = datetime.now().strftime("%Y%m%d_%H%M%S")
    generated_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    ws["B3"] = old_version
    ws["B2"] = new_version
    ws["B4"] = generated_at
    ws["B5"] = "current"

    return new_version

def build_output_filenames(shift_sheet_name: str, version: str):
    cleaning_filename = f"{shift_sheet_name}_掃除当番表.xlsx"
    master_filename = f"{shift_sheet_name}_更新済みマスタ_{version}.xlsx"
    zip_filename = f"{shift_sheet_name}_出力一式_{version}.zip"
    return cleaning_filename, master_filename, zip_filename

def prepare_cleaning_generation(
    shift_path: str,
    master_path: str,
    shift_sheet_name: str,
    clean_days: str,
    order_sheet_name: str,
    state_sheet_name: str,
):
    # シフト表読み込み
    shift_wb = openpyxl.load_workbook(shift_path)
    if shift_sheet_name not in shift_wb.sheetnames:
        raise HTTPException(
            status_code=400,
            detail=f"シフト表ファイルに '{shift_sheet_name}' シートが見つかりません"
        )

    shift_ws = shift_wb[shift_sheet_name]
    parsed_shift = parse_shift_table_for_cleaning(shift_ws, shift_sheet_name)

    # マスタファイル読み込み
    master_wb = openpyxl.load_workbook(master_path)

    # 版管理チェック（なければ初回作成）
    version_info = validate_or_initialize_version_sheet(master_wb)

    if order_sheet_name not in master_wb.sheetnames:
        raise HTTPException(
            status_code=400,
            detail=f"マスタファイルに '{order_sheet_name}' シートが見つかりません"
        )

    order_ws = master_wb[order_sheet_name]
    order_data = parse_order_sheet(order_ws)

    state_ws = ensure_state_sheet(master_wb, state_sheet_name, order_data.keys())
    state_data = read_state_sheet(state_ws, order_data)

    clean_days_normalized = (
        (clean_days or "")
        .replace("　", "")
        .replace(" ", "")
        .replace("・", "")
        .strip()
    )
    if not clean_days_normalized:
        clean_days_normalized = "月水金"

    assignments, new_state = assign_cleaning_with_state(
        parsed_shift,
        order_data,
        state_data,
        clean_days_normalized
    )

    m = re.match(r"(\d{8})-(\d{8})", shift_sheet_name)
    if not m:
        raise HTTPException(status_code=400, detail="shift_sheet_name の形式が不正です")

    period_start = datetime.strptime(m.group(1), "%Y%m%d").date()
    period_end = datetime.strptime(m.group(2), "%Y%m%d").date()
    period_label = f"{period_start.year}/{period_start.month}/{period_start.day}〜{period_end.year}/{period_end.month}/{period_end.day}"

    return {
        "master_wb": master_wb,
        "order_ws": order_ws,
        "state_ws": state_ws,
        "order_data": order_data,
        "assignments": assignments,
        "new_state": new_state,
        "period_label": period_label,
        "shift_sheet_name": shift_sheet_name,
        "version_info": version_info,
    }

# =========================
# API
# =========================
@app.get("/health")
def health_check():
    return {"status": "ok"}

@app.post("/generate-cleaning-sheet")
async def generate_cleaning_sheet(
    shift_file: UploadFile = File(...),
    master_file: UploadFile = File(...),
    shift_sheet_name: str = Form(...),
    clean_days: str = Form("月水金"),
    order_sheet_name: str = Form(DEFAULT_ORDER_SHEET_NAME),
    state_sheet_name: str = Form(DEFAULT_STATE_SHEET_NAME),
):
    if not shift_sheet_name.strip():
        raise HTTPException(status_code=400, detail="shift_sheet_name is required")

    shift_filename = shift_file.filename or "shift_file.xlsx"
    shift_ext = os.path.splitext(shift_filename)[1].lower()
    if shift_ext not in [".xlsx", ".xlsm", ".xltx", ".xltm"]:
        raise HTTPException(status_code=400, detail=f"Unsupported shift file type: {shift_ext}")

    master_filename = master_file.filename or "master_file.xlsx"
    master_ext = os.path.splitext(master_filename)[1].lower()
    if master_ext not in [".xlsx", ".xlsm", ".xltx", ".xltm"]:
        raise HTTPException(status_code=400, detail=f"Unsupported master file type: {master_ext}")

    try:
        with tempfile.NamedTemporaryFile(delete=False, suffix=shift_ext) as tmp_shift:
            tmp_shift.write(await shift_file.read())
            shift_path = tmp_shift.name

        with tempfile.NamedTemporaryFile(delete=False, suffix=master_ext) as tmp_master:
            tmp_master.write(await master_file.read())
            master_path = tmp_master.name

        prepared = prepare_cleaning_generation(
            shift_path=shift_path,
            master_path=master_path,
            shift_sheet_name=shift_sheet_name,
            clean_days=clean_days,
            order_sheet_name=order_sheet_name,
            state_sheet_name=state_sheet_name,
        )

        output_wb = Workbook()
        default_ws = output_wb.active
        output_wb.remove(default_ws)

        output_sheet_name = f"掃除当番_{prepared['shift_sheet_name']}"
        write_cleaning_sheet(
            output_wb,
            output_sheet_name,
            prepared["assignments"],
            prepared["period_label"]
        )

        with NamedTemporaryFile(delete=False, suffix=".xlsx") as out_tmp:
            output_path = out_tmp.name

        output_wb.save(output_path)

        return FileResponse(
            output_path,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename=f"{prepared['shift_sheet_name']}_掃除当番表.xlsx"
        )

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Excel read error: {e}")

@app.post("/generate-updated-master")
async def generate_updated_master(
    shift_file: UploadFile = File(...),
    master_file: UploadFile = File(...),
    shift_sheet_name: str = Form(...),
    clean_days: str = Form("月水金"),
    order_sheet_name: str = Form(DEFAULT_ORDER_SHEET_NAME),
    state_sheet_name: str = Form(DEFAULT_STATE_SHEET_NAME),
):
    if not shift_sheet_name.strip():
        raise HTTPException(status_code=400, detail="shift_sheet_name is required")

    shift_filename = shift_file.filename or "shift_file.xlsx"
    shift_ext = os.path.splitext(shift_filename)[1].lower()
    if shift_ext not in [".xlsx", ".xlsm", ".xltx", ".xltm"]:
        raise HTTPException(status_code=400, detail=f"Unsupported shift file type: {shift_ext}")

    master_filename = master_file.filename or "master_file.xlsx"
    master_ext = os.path.splitext(master_filename)[1].lower()
    if master_ext not in [".xlsx", ".xlsm", ".xltx", ".xltm"]:
        raise HTTPException(status_code=400, detail=f"Unsupported master file type: {master_ext}")

    try:
        with tempfile.NamedTemporaryFile(delete=False, suffix=shift_ext) as tmp_shift:
            tmp_shift.write(await shift_file.read())
            shift_path = tmp_shift.name

        with tempfile.NamedTemporaryFile(delete=False, suffix=master_ext) as tmp_master:
            tmp_master.write(await master_file.read())
            master_path = tmp_master.name

        prepared = prepare_cleaning_generation(
            shift_path=shift_path,
            master_path=master_path,
            shift_sheet_name=shift_sheet_name,
            clean_days=clean_days,
            order_sheet_name=order_sheet_name,
            state_sheet_name=state_sheet_name,
        )

        master_wb = prepared["master_wb"]
        order_ws = prepared["order_ws"]
        state_ws = prepared["state_ws"]
        order_data = prepared["order_data"]
        new_state = prepared["new_state"]

        # マスタ更新
        write_state_sheet(state_ws, new_state)
        update_order_sheet_yellow(order_ws, order_data, new_state)
        new_version = update_version_info(master_wb)

        with NamedTemporaryFile(delete=False, suffix=".xlsx") as master_tmp:
            updated_master_path = master_tmp.name

        master_wb.save(updated_master_path)

        return FileResponse(
            updated_master_path,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename=f"{prepared['shift_sheet_name']}_更新済みマスタ_{new_version}.xlsx"
        )

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Excel read error: {e}")

@app.post("/generate-cleaning-schedule")
async def generate_cleaning_schedule(
    shift_file: UploadFile = File(...),
    master_file: UploadFile = File(...),
    shift_sheet_name: str = Form(...),
    clean_days: str = Form("月水金"),
    order_sheet_name: str = Form(DEFAULT_ORDER_SHEET_NAME),
    state_sheet_name: str = Form(DEFAULT_STATE_SHEET_NAME),
    target_period_label: Optional[str] = Form(None),
):
    if not shift_sheet_name.strip():
        raise HTTPException(status_code=400, detail="shift_sheet_name is required")

    shift_filename = shift_file.filename or "shift_file.xlsx"
    shift_ext = os.path.splitext(shift_filename)[1].lower()
    if shift_ext not in [".xlsx", ".xlsm", ".xltx", ".xltm"]:
        raise HTTPException(status_code=400, detail=f"Unsupported shift file type: {shift_ext}")

    master_filename = master_file.filename or "master_file.xlsx"
    master_ext = os.path.splitext(master_filename)[1].lower()
    if master_ext not in [".xlsx", ".xlsm", ".xltx", ".xltm"]:
        raise HTTPException(status_code=400, detail=f"Unsupported master file type: {master_ext}")

    try:
        # ① シフト表ファイル保存
        with tempfile.NamedTemporaryFile(delete=False, suffix=shift_ext) as tmp_shift:
            tmp_shift.write(await shift_file.read())
            shift_path = tmp_shift.name

        # ② マスタファイル保存
        with tempfile.NamedTemporaryFile(delete=False, suffix=master_ext) as tmp_master:
            tmp_master.write(await master_file.read())
            master_path = tmp_master.name

        # シフト表読み込み
        shift_wb = openpyxl.load_workbook(shift_path)
        if shift_sheet_name not in shift_wb.sheetnames:
            raise HTTPException(
                status_code=400,
                detail=f"シフト表ファイルに '{shift_sheet_name}' シートが見つかりません"
            )

        shift_ws = shift_wb[shift_sheet_name]
        parsed_shift = parse_shift_table_for_cleaning(shift_ws, shift_sheet_name)

        # マスタファイル読み込み
        master_wb = openpyxl.load_workbook(master_path)

        # 版管理チェック（なければ初回作成）
        version_info = validate_or_initialize_version_sheet(master_wb)

        if order_sheet_name not in master_wb.sheetnames:
            raise HTTPException(
                status_code=400,
                detail=f"マスタファイルに '{order_sheet_name}' シートが見つかりません"
            )

        order_ws = master_wb[order_sheet_name]
        order_data = parse_order_sheet(order_ws)

        state_ws = ensure_state_sheet(master_wb, state_sheet_name, order_data.keys())
        state_data = read_state_sheet(state_ws, order_data)

        clean_days_normalized = (
            (clean_days or "")
            .replace("　", "")
            .replace(" ", "")
            .replace("・", "")
            .strip()
        )
        if not clean_days_normalized:
            clean_days_normalized = "月水金"

        assignments, new_state = assign_cleaning_with_state(
            parsed_shift,
            order_data,
            state_data,
            clean_days_normalized
        )

        # 状態管理シート更新（マスタ側）
        write_state_sheet(state_ws, new_state)

        # 順番表シートの黄色更新（マスタ側）
        update_order_sheet_yellow(order_ws, order_data, new_state)

        # 版管理更新
        new_version = update_version_info(master_wb)

        # -------------------------
        # ファイル名を決定
        # -------------------------
        cleaning_filename, master_filename, zip_filename = build_output_filenames(
            shift_sheet_name,
            new_version
        )

        # -------------------------
        # 1. 更新済みマスタを保存
        # -------------------------
        with NamedTemporaryFile(delete=False, suffix=".xlsx") as master_tmp:
            updated_master_path = master_tmp.name

        master_wb.save(updated_master_path)

        # -------------------------
        # 2. 掃除当番表だけ別Excelで新規作成
        # -------------------------
        output_wb = Workbook()
        default_ws = output_wb.active
        output_wb.remove(default_ws)

        m = re.match(r"(\d{8})-(\d{8})", shift_sheet_name)
        if not m:
            raise HTTPException(status_code=400, detail="shift_sheet_name の形式が不正です")

        period_start = datetime.strptime(m.group(1), "%Y%m%d").date()
        period_end = datetime.strptime(m.group(2), "%Y%m%d").date()
        period_label = f"{period_start.year}/{period_start.month}/{period_start.day}〜{period_end.year}/{period_end.month}/{period_end.day}"

        output_sheet_name = f"掃除当番_{shift_sheet_name}"
        write_cleaning_sheet(output_wb, output_sheet_name, assignments, period_label)

        with NamedTemporaryFile(delete=False, suffix=".xlsx") as cleaning_tmp:
            cleaning_output_path = cleaning_tmp.name

        output_wb.save(cleaning_output_path)

        # -------------------------
        # 3. ZIP にまとめる
        # -------------------------
        with NamedTemporaryFile(delete=False, suffix=".zip") as zip_tmp:
            zip_output_path = zip_tmp.name

        with zipfile.ZipFile(zip_output_path, "w", zipfile.ZIP_DEFLATED) as zf:
            zf.write(cleaning_output_path, arcname=cleaning_filename)
            zf.write(updated_master_path, arcname=master_filename)

        return FileResponse(
            zip_output_path,
            media_type="application/zip",
            filename=zip_filename
        )
    
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Excel read error: {e}")